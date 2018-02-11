"""
FTP上下载的高空分钟数据进行格式转换后的txt文件提取各高度层风向风速
保存的风向风速数据均为float类型
"""

import os
import datetime
import re


# 文件名（包括路径），所有规定层高度，需要的数据在每一行的位置
def read_uperfile(uperfile, uper_high):
    # 经纬度偏差分别所处的位置
    wdIndex = 14
    wsIndex = 15
    wdData = []
    wsData = []
    i = 0
    with open(uperfile, 'r') as rf:
        for line in rf:
            buffer = line.split()
            # 判断是否到达相应的规定层,若到达，则取出在uperdata_index处的数据
            # uperdata_index——风向：14，风速：15
            if i < len(uper_high):
                daDif = float(buffer[16]) - uper_high[i] * 1000
                if daDif > 0:
                    wdData.append(buffer[wdIndex])
                    wsData.append(buffer[wsIndex])
                    i += 1
            else:
                break
    return wdData, wsData


# 将提取的风向风速保存至excel
def save_uperdata(filename, head, data1, data2):
    import openpyxl
    wb = openpyxl.Workbook()
    sheet1 = wb.create_sheet(title='风向', index=0)
    sheet1.cell(row=1, column=1).value = '年'
    sheet1.cell(row=1, column=2).value = '月'
    sheet1.cell(row=1, column=3).value = '日'
    sheet1.cell(row=1, column=4).value = '时'

    sheet2 = wb.create_sheet(title='风速', index=1)
    sheet2.cell(row=1, column=1).value = '年'
    sheet2.cell(row=1, column=2).value = '月'
    sheet2.cell(row=1, column=3).value = '日'
    sheet2.cell(row=1, column=4).value = '时'

    for head_index, head_data in enumerate(head):
        sheet1.cell(row=1, column=head_index + 5).value = head_data
        sheet2.cell(row=1, column=head_index + 5).value = head_data
    data1_row = 2
    data2_row = 2
    for datime1, data1s in data1.items():
        # 将时间拆分为年、月、日、时
        dateTime1 = re.findall(r'[\w]+', datime1)
        sheet1.cell(row=data1_row, column=1).value = int(dateTime1[0])
        sheet1.cell(row=data1_row, column=2).value = int(dateTime1[1])
        sheet1.cell(row=data1_row, column=3).value = int(dateTime1[2])
        sheet1.cell(row=data1_row, column=4).value = int(dateTime1[3])
        for j, p in enumerate(data1s):
            sheet1.cell(row=data1_row, column=j + 5).value = float(p)
        data1_row += 1

    for datime2, data2s in data2.items():
        dateTime2 = re.findall(r'[\w]+', datime2)
        sheet2.cell(row=data2_row, column=1).value = int(dateTime2[0])
        sheet2.cell(row=data2_row, column=2).value = int(dateTime2[1])
        sheet2.cell(row=data2_row, column=3).value = int(dateTime2[2])
        sheet2.cell(row=data2_row, column=4).value = int(dateTime2[3])
        for j, p in enumerate(data2s):
            sheet2.cell(row=data2_row, column=j + 5).value = float(p)
        data2_row += 1

    wb.save(filename)


# 定义excel第一行头部
uperHigh = [0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 5.5, 6.0, 7.0,
            8.0, 9.0, 10.0, 10.5, 12.0, 14.0, 16.0, 18.0, 20.0,
            22.0, 24.0, 26.0, 28.0, 30.0, 32.0, 34.0, 36.0, 38.0, 40.0]
Heads = [str(uperHigh[i]) + 'km' for i in range(len(uperHigh))]
# 需要提取的文本文件位置及文件名
root = r'G:\python\uper_hb'
basePaths = os.listdir(root)
for ph in basePaths:
    basePath = os.path.join(root, ph)
    fileName = [file for file in os.listdir(basePath) if file.endswith('.txt')]
    # 定义文件保存位置及文件名
    endFile = basePath + '.xlsx'
    # 定义风向风速
    wdS = {}
    wsS = {}
    for file in fileName:
        myFile = os.path.join(basePath, file)
        # 从文件名中取出日期、时间,并将其转为北京时
        dateTime = datetime.datetime.strptime(re.findall(r'[\w]+', os.path.basename(myFile))[1], '%Y%m%d%H')
        dateTime = (dateTime + datetime.timedelta(hours=8)).strftime('%Y-%m-%d %H')
        # 提取风向风速
        wd, ws = read_uperfile(myFile, uperHigh)
        wdS[dateTime] = wd
        wsS[dateTime] = ws
    save_uperdata(endFile, Heads, wdS, wsS)
