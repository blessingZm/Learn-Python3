import openpyxl
import os


def get_users():
    filepath = os.path.abspath('../FTPsome')
    filename = '省级AJ文件FTP访问用户名、密码.xlsx'
    wb = openpyxl.load_workbook(os.path.join(filepath, filename))
    sheet = wb.active
    users = []
    for row in range(3, sheet.max_row + 1):
        userCity = sheet.cell(row=row, column=1).value
        userName = sheet.cell(row=row, column=2).value
        userPass = sheet.cell(row=row, column=3).value
        users.append((userCity, userName, userPass))
    return users

