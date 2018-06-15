"""
FTP上下载的高空分钟数据进行格式转换后的txt文件提取各高度层数据
保存的数据均为float类型
"""

import os
import datetime
import re
import openpyxl


# 文件名（包括路径），所有规定层高度，需要的数据在每一行的位置
def read_uperfile(uperfile, uper_high, data_index):
    meter_datas = []
    i = 0
    with open(uperfile, 'r') as rf:
        for line in rf:
            buffer = line.split()
            # 判断是否到达相应的规定层,若到达，则取出在uperdata_index处的数据
            if i < len(uper_high):
                da_dif = float(buffer[HIGH_INDEX]) - uper_high[i] * 1000
                if da_dif >= 0:
                    if float(buffer[data_index]) < 999:
                        meter_datas.append(buffer[data_index])
                        i += 1
                    elif i < len(uper_high) - 1 and float(buffer[HIGH_INDEX]) - uper_high[i + 1] * 1000 >= 0:
                        meter_datas.append(' ')
                        i += 1
                    else:
                        continue
            else:
                break
    return meter_datas


# 将提取的数据保存至excel
def save_uperdata(filename, head, data, sheet_name, index_num):
    wb = openpyxl.load_workbook(filename=filename)
    sheet = wb.create_sheet(title=sheet_name, index=index_num)
    sheet.cell(row=1, column=1).value = '年'
    sheet.cell(row=1, column=2).value = '月'
    sheet.cell(row=1, column=3).value = '日'
    sheet.cell(row=1, column=4).value = '时'

    for head_index, head_data in enumerate(head):
        sheet.cell(row=1, column=head_index + 5).value = head_data
    data_row = 2
    for datime, datas in data.items():
        # 将时间拆分为年、月、日、时
        date_time = re.findall(r'[\w]+', datime)
        sheet.cell(row=data_row, column=1).value = int(date_time[0])
        sheet.cell(row=data_row, column=2).value = int(date_time[1])
        sheet.cell(row=data_row, column=3).value = int(date_time[2])
        sheet.cell(row=data_row, column=4).value = int(date_time[3])
        for j, p in enumerate(datas):
            sheet.cell(row=data_row, column=j + 5).value = float(p)
        data_row += 1
    wb.save(filename)


if __name__ == "__main__":
    # 位势高度在数据行所处的位置
    HIGH_INDEX = 16
    # 高空气象要素及其对应在每行的位置
    METER_NAMES = ['气温', '气压', '相对湿度', '风向', '风速']
    METER_INDEX = [6, 7, 8, 14, 15]
    # 定义excel第一行头部
    uperHigh = [0.5, 1.0, 1.5, 2.0, 3.0, 4.0, 5.0, 5.5, 6.0, 7.0,
                8.0, 9.0, 10.0, 10.5, 12.0, 14.0, 16.0, 18.0, 20.0,
                22.0, 24.0, 26.0, 28.0, 30.0, 32.0, 34.0, 36.0, 38.0, 40.0]
    Heads = [str(uperHigh[i]) + 'km' for i in range(len(uperHigh))]
    # 需要提取的文本文件位置及文件名
    root = r'G:\python\uper_hb'
    basePaths = os.listdir(root)
    for ph in basePaths:
        if os.path.isdir(os.path.join(root, ph)):
            basePath = os.path.join(root, ph)
            fileName = [file for file in os.listdir(basePath) if file.endswith('.txt')]
            # 定义文件保存位置及文件名
            endFile = os.path.join(root, ph + '.xlsx')
            # 判断最终的文件是否存在，若不存在则创建
            if not os.path.exists(endFile):
                wb = openpyxl.Workbook()
                wb.save(endFile)
            # 开始提取各气象要素并保存
            for i, meterName in enumerate(METER_NAMES):
                endDatas = {}
                for file in fileName:
                    myFile = os.path.join(basePath, file)
                    # 从文件名中取出日期、时间,并将其转为北京时
                    dateTime = datetime.datetime.strptime(re.findall(r'[\w]+', os.path.basename(myFile))[1], '%Y%m%d%H')
                    dateTime = (dateTime + datetime.timedelta(hours=8)).strftime('%Y-%m-%d %H')

                    meterDatas = read_uperfile(myFile, uperHigh, METER_INDEX[i])
                    endDatas[dateTime] = meterDatas
                save_uperdata(endFile, Heads, endDatas, sheet_name=meterName, index_num=i)
